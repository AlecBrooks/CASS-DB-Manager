import os
import sys
import math
import psutil
import logging
import argparse
import sqlite3
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt='%Y-%m-%d %H:%M:%S'
)

def load_config_file(config_path: str) -> dict:
    if not os.path.exists(config_path):
        logging.error(f"Config file not found: {config_path}")
        sys.exit(1)

    cfg = {}
    with open(config_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                key, val = key.strip(), val.strip()
                cfg[key] = val
    return cfg

class CassSpeciation:
    def __init__(self,
                 start_date: datetime,
                 end_date: datetime,
                 constants_cfg: dict,
                 db_cfg: dict,
                 avg_interval_seconds: int):

        required_bc = ["BC1", "BC2", "BC3", "BC4", "BC5", "BC6", "BC7"]
        required_keys = [
            "AAE_bb", "AAE_ff", "AAE_bc",
            "MAC_bb", "MAC_ff",
            "POA_POC_Ratio", "SOA_SOC_Ratio",
            "MAC_BrC_Prim", "MAC_BrC_Sec",
            "Time_Delta"
        ]

        missing_bc_keys = [k for k in required_bc if k not in constants_cfg]
        missing_const_keys = [k for k in required_keys if k not in constants_cfg]

        if missing_bc_keys or missing_const_keys:
            if missing_bc_keys:
                logging.error(f"Missing BC multiplier keys in constants.conf: {missing_bc_keys}")
            if missing_const_keys:
                logging.error(f"Missing constant keys in constants.conf: {missing_const_keys}")
            sys.exit(1)

        self.bc_multipliers = {
            "BC1": float(constants_cfg["BC1"]) / 1000.0,
            "BC2": float(constants_cfg["BC2"]) / 1000.0,
            "BC3": float(constants_cfg["BC3"]) / 1000.0,
            "BC4": float(constants_cfg["BC4"]) / 1000.0,
            "BC5": float(constants_cfg["BC5"]) / 1000.0,
            "BC6": float(constants_cfg["BC6"]) / 1000.0,
            "BC7": float(constants_cfg["BC7"]) / 1000.0,
        }

        self.AAE_bb = float(constants_cfg["AAE_bb"])
        self.AAE_ff = float(constants_cfg["AAE_ff"])
        self.AAE_bc = float(constants_cfg["AAE_bc"])
        self.MAC_bb = float(constants_cfg["MAC_bb"])
        self.MAC_ff = float(constants_cfg["MAC_ff"])
        self.POA_POC_Ratio = float(constants_cfg["POA_POC_Ratio"])
        self.SOA_SOC_Ratio = float(constants_cfg["SOA_SOC_Ratio"])
        self.MAC_BrC_Prim = float(constants_cfg["MAC_BrC_Prim"])
        self.MAC_BrC_Sec = float(constants_cfg["MAC_BrC_Sec"])
        self.Time_Delta = int(float(constants_cfg["Time_Delta"]))
        self.avg_interval_seconds = avg_interval_seconds

        # Updated required keys for SQLite configuration
        needed_db_keys = ["dbPath", "AE33_Table", "TCA_Table"]
        for k in needed_db_keys:
            if k not in db_cfg:
                logging.error(f"Missing key '{k}' in db.conf")
                sys.exit(1)

        self.db_path = db_cfg["dbPath"]
        self.ae33_table = db_cfg["AE33_Table"]
        self.tca_table = db_cfg["TCA_Table"]

        self.TCA_AE_hourly = pd.DataFrame()

        days_in_range = (end_date - start_date).days + 1
        remainder = days_in_range % self.Time_Delta
        if remainder != 0:
            add_days = self.Time_Delta - remainder
            end_date = end_date + timedelta(days=add_days)
            logging.info(f"Extending end date by {add_days} day(s) to get uniform chunking. New end: {end_date.date()}")

        self.start_date = start_date
        self.end_date = end_date

        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        out_root_dir = os.path.join(root_dir, "data", "CASSOutput")
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.run_dir = os.path.join(out_root_dir, timestamp_str)
        os.makedirs(self.run_dir, exist_ok=True)

        self.plot_dir = os.path.join(self.run_dir, "plots")
        os.makedirs(self.plot_dir, exist_ok=True)
        
        self.rsquared_dir = os.path.join(self.run_dir, "RSquared")
        os.makedirs(self.rsquared_dir, exist_ok=True)

        self.output_xlsx = os.path.join(self.run_dir, "CASSOutput.xlsx")

    def run(self):
        logging.info("Starting CassSpeciation run.")

        if self.is_excel_open(self.output_xlsx):
            logging.error(f"Excel file is open: {self.output_xlsx}. Please close it first.")
            sys.exit(1)

        margin_start = self.start_date - timedelta(days=self.Time_Delta)
        margin_end = self.end_date + timedelta(days=self.Time_Delta)
        logging.info("Fetching hourly data from DB.")
        
        df = self.fetch_hourly_data(margin_start, margin_end)
        if df.empty:
            logging.warning("No rows returned from DB for the requested date range (with margin).")
            return

        logging.info(f"Fetched {len(df)} records from DB.")

        logging.info("Renaming columns and applying multipliers to BC data.")
        df.rename(columns={
            'avg_TConc': 'TCconc',
            'avg_CO2': 'CO2',
            'avg_EC': 'EC',
            'avg_OC': 'OC',
            'avg_AE33_BC6': 'AE33_BC6',
            'avg_BC1': 'B-abs1',
            'avg_BC2': 'B-abs2',
            'avg_BC3': 'B-abs3',
            'avg_BC4': 'B-abs4',
            'avg_BC5': 'B-abs5',
            'avg_BC6': 'B-abs6',
            'avg_BC7': 'B-abs7'
        }, inplace=True)

        # Convert the 'Date_and_Time' column to datetime objects
        df['Date_and_Time'] = pd.to_datetime(df['Date_and_Time'])

        for i in range(1, 8):
            key = f"B-abs{i}"
            df[key] = df[key] * self.bc_multipliers[f"BC{i}"]

        self.TCA_AE_hourly = df.copy()

        logging.info("Starting chunk-based correlation and final calculations.")
        self.calculate_final_columns()
        logging.info("Finished correlation and calculations.")

        self.TCA_AE_hourly['B-abs6_Val'] = self.TCA_AE_hourly['AE33_BC6'] * 7.77 / 1000

        final_df = self.TCA_AE_hourly[
            (self.TCA_AE_hourly['Date_and_Time'] >= self.start_date) &
            (self.TCA_AE_hourly['Date_and_Time'] <= self.end_date)
        ].copy()

        final_df.fillna("NA", inplace=True)

        for c_ in ["date", "hour"]:
            if c_ in final_df.columns:
                final_df.drop(columns=[c_], inplace=True)

        ordered_columns = [
            "Date_and_Time", "B-abs1", "B-abs2", "B-abs3", "B-abs4", "B-abs5", "B-abs6", "B-abs7",
            "TCconc", "CO2", "EC", "OC", "AE33_BC6", "B-abs6_Val", "B-abs-ff", "B-abs-bb",
            "BC-ff", "BC-bb", "B-abs-BC", "B-abs-Brc", "BrC", "BrC-abs-Sec", "SOC", "POC",
            "BrC-abs-Prim", "POA", "SOA", "POA_BrC", "SOA_BrC", "POA_WtC", "SOA_WtC"
        ]
        final_df = final_df[ordered_columns]

        logging.info("Writing main data to Excel.")
        self.write_to_excel(final_df, sheet_name="data")

        logging.info("Writing constants to Excel.")
        self.write_constants_sheet()

        logging.info("Fetching and writing gap data for TCA and AE33.")
        df_tca_gaps = self.fetch_tca_gaps(self.start_date, self.end_date)
        if not df_tca_gaps.empty:
            df_tca_gaps.fillna("NA", inplace=True)
        self.write_to_excel(df_tca_gaps, sheet_name="TCA Gaps")

        df_ae33_gaps = self.fetch_ae33_gaps(self.start_date, self.end_date)
        if not df_ae33_gaps.empty:
            df_ae33_gaps.fillna("NA", inplace=True)
        self.write_to_excel(df_ae33_gaps, sheet_name="AE33 Gaps")

        logging.info("Generating main plots (time series and diurnal).")
        self.produce_main_plots("Speciation")

        logging.info("CassSpeciation run completed successfully.")

    def fetch_ae33_gaps(self, start_: datetime, end_: datetime) -> pd.DataFrame:
        sql = f"""
WITH ordered_timestamps AS (
    SELECT 
        datetime,
        LAG(datetime) OVER (ORDER BY datetime) AS prev_timestamp
    FROM {self.ae33_table}
    WHERE datetime >= '{start_.strftime("%Y-%m-%d")}'
      AND datetime <= '{end_.strftime("%Y-%m-%d")}'
)
SELECT 
    prev_timestamp AS gap_start,
    datetime AS gap_end,
    ((strftime('%s', datetime) - strftime('%s', prev_timestamp)) / 60) as minute_duration
FROM ordered_timestamps
WHERE ((strftime('%s', datetime) - strftime('%s', prev_timestamp)) / 60) > 1
ORDER BY prev_timestamp;
"""
        cnx = sqlite3.connect(self.db_path)
        df = pd.read_sql_query(sql, cnx)
        cnx.close()
        logging.info("Fetched AE33 gap data.")
        return df

    def fetch_tca_gaps(self, start_: datetime, end_: datetime) -> pd.DataFrame:
        sql = f"""
WITH ordered_data AS (
    SELECT 
        StartTimeLocal,
        LAG(StartTimeLocal) OVER (ORDER BY StartTimeLocal) AS prev_timestamp
    FROM {self.tca_table}
    WHERE StartTimeLocal >= '{start_.strftime("%Y-%m-%d")}'
      AND StartTimeLocal <= '{end_.strftime("%Y-%m-%d")}'
)
SELECT 
    prev_timestamp AS gap_start,
    StartTimeLocal AS gap_end,
    ROUND(((strftime('%s', StartTimeLocal) - strftime('%s', prev_timestamp)) / 60.0) / 60.0, 2) AS gap_duration_hours
FROM ordered_data
WHERE ((strftime('%s', StartTimeLocal) - strftime('%s', prev_timestamp)) / 60) > 60
ORDER BY prev_timestamp;
"""
        cnx = sqlite3.connect(self.db_path)
        df = pd.read_sql_query(sql, cnx)
        cnx.close()
        logging.info("Fetched TCA gap data.")
        return df

    def fetch_hourly_data(self, start_: datetime, end_: datetime) -> pd.DataFrame:
        interval = self.avg_interval_seconds
        sql = f"""
WITH t AS (
    SELECT 
      datetime((CAST(strftime('%s', StartTimeLocal) AS INTEGER) / {interval}) * {interval}, 'unixepoch') as bucket,
      AVG(TCconc)    AS avg_TConc,
      AVG(CO2)       AS avg_CO2,
      AVG(EC)        AS avg_EC,
      AVG(OC)        AS avg_OC,
      AVG(AE33_BC6)  AS avg_AE33_BC6
   FROM {self.tca_table}
   WHERE StartTimeLocal >= '{start_.strftime("%Y-%m-%d")}'
     AND StartTimeLocal <  '{(end_ + timedelta(days=1)).strftime("%Y-%m-%d")}'
   GROUP BY bucket
), a AS (
   SELECT 
      datetime((CAST(strftime('%s', datetime(date || ' ' || time)) AS INTEGER) / {interval}) * {interval}, 'unixepoch') as bucket,
      AVG(BC1) AS avg_BC1, 
      AVG(BC2) AS avg_BC2, 
      AVG(BC3) AS avg_BC3, 
      AVG(BC4) AS avg_BC4, 
      AVG(BC5) AS avg_BC5, 
      AVG(BC6) AS avg_BC6, 
      AVG(BC7) AS avg_BC7
   FROM {self.ae33_table}
   WHERE date >= '{start_.strftime("%Y-%m-%d")}'
     AND date <  '{(end_ + timedelta(days=1)).strftime("%Y-%m-%d")}'
   GROUP BY bucket
), buckets AS (
    SELECT bucket FROM t
    UNION
    SELECT bucket FROM a
)
SELECT 
    b.bucket as Date_and_Time,
    COALESCE(t.avg_TConc, -99) as avg_TConc,
    COALESCE(t.avg_CO2, -99) as avg_CO2,
    COALESCE(t.avg_EC, -99) as avg_EC,
    COALESCE(t.avg_OC, -99) as avg_OC,
    COALESCE(t.avg_AE33_BC6, -99) as avg_AE33_BC6,
    COALESCE(a.avg_BC1, -99) as avg_BC1,
    COALESCE(a.avg_BC2, -99) as avg_BC2,
    COALESCE(a.avg_BC3, -99) as avg_BC3,
    COALESCE(a.avg_BC4, -99) as avg_BC4,
    COALESCE(a.avg_BC5, -99) as avg_BC5,
    COALESCE(a.avg_BC6, -99) as avg_BC6,
    COALESCE(a.avg_BC7, -99) as avg_BC7
FROM buckets b
LEFT JOIN t ON b.bucket = t.bucket
LEFT JOIN a ON b.bucket = a.bucket
ORDER BY b.bucket;
"""
        cnx = sqlite3.connect(self.db_path)
        df = pd.read_sql_query(sql, cnx)
        cnx.close()
        return df

    def is_excel_open(self, file_path: str) -> bool:
        for proc in psutil.process_iter(['pid', 'name']):
            if 'EXCEL.EXE' in proc.info.get('name', ''):
                try:
                    for ofile in proc.open_files():
                        if ofile.path.lower() == file_path.lower():
                            return True
                except psutil.AccessDenied:
                    pass
        return False

    def calculate_final_columns(self):
        self.min_r2_calculation_for_brC()
        self.min_r2_calculation_for_SOC()

        df = self.TCA_AE_hourly
        numFF = df['B-abs7'] - df['B-abs2'] * (950/470)**(-self.AAE_bb)
        denFF = ((950/470)**(-self.AAE_ff) - (950/470)**(-self.AAE_bb))
        df['B-abs-ff'] = numFF / denFF

        numBB = df['B-abs7'] - df['B-abs2'] * (950/470)**(-self.AAE_ff)
        denBB = ((950/470)**(-self.AAE_ff) - (950/470)**(-self.AAE_bb))
        df['B-abs-bb'] = numBB / denBB

        part2 = self.MAC_ff / self.MAC_bb
        part3_nm = 1 - (df['B-abs2']/df['B-abs7'])*(950/470)**(-self.AAE_ff)
        part3_dm = 1 - (df['B-abs2']/df['B-abs7'])*(950/470)**(-self.AAE_bb)
        BCff_BC_Ratio = 1 / (1 - part2*(part3_nm/part3_dm))
        df['BC-ff'] = df['AE33_BC6'] * BCff_BC_Ratio
        df['BC-bb'] = df['AE33_BC6'] - df['BC-ff']

        df['B-abs-BC'] = df['B-abs6']*(470/880)**(-self.AAE_bc)
        df['B-abs-Brc'] = df['B-abs2'] - df['B-abs-BC']
        df['BrC'] = df['OC'] - df['AE33_BC6']

        df['POC'] = df['OC'] - df['SOC']
        df['POA'] = df['POC'] * self.POA_POC_Ratio
        df['SOA'] = df['SOC'] * self.SOA_SOC_Ratio
        df['BrC-abs-Prim'] = df['B-abs-Brc'] - df['BrC-abs-Sec']

        df['POA_BrC'] = df['BrC-abs-Prim'] / self.MAC_BrC_Prim
        df['SOA_BrC'] = df['BrC-abs-Sec']  / self.MAC_BrC_Sec
        df['POA_WtC'] = df['POA'] - df['POA_BrC']
        df['SOA_WtC'] = df['SOA'] - df['SOA_BrC']

        for idx, row in df.iterrows():
            if (row['B-abs7'] == -99) or (row['B-abs2'] == -99):
                df.at[idx, 'B-abs-bb'] = -99
                df.at[idx, 'B-abs-ff'] = -99
                df.at[idx, 'BC-ff'] = -99
                df.at[idx, 'BC-bb'] = -99
                df.at[idx, 'B-abs-Brc'] = -99
                df.at[idx, 'B-abs-BC'] = -99
                df.at[idx, 'BrC-abs-Sec'] = -99

            if pd.isna(row['TCconc']) or (row['TCconc'] == -99):
                for col_ in [
                    'SOC', 'POC', 'SOA', 'POA', 'AE33_BC6', 'BrC', 'BrC-abs-Sec',
                    'BrC-abs-Prim', 'POA_BrC', 'SOA_BrC', 'POA_WtC', 'SOA_WtC',
                    'TCconc', 'CO2', 'EC', 'OC', 'BC-ff', 'BC-bb'
                ]:
                    df.at[idx, col_] = -99

    def min_r2_calculation_for_brC(self):
        df = self.TCA_AE_hourly
        min_dt = df['Date_and_Time'].min()
        max_dt = df['Date_and_Time'].max()

        found_3day = False
        start_ = min_dt
        while start_ < max_dt:
            chunk_end = start_ + timedelta(days=self.Time_Delta)
            chunk = df[(df['Date_and_Time'] >= start_) & (df['Date_and_Time'] < chunk_end)]

            unique_days_in_chunk = chunk['Date_and_Time'].dt.date.nunique()
            if unique_days_in_chunk < 3:
                start_ = chunk_end
                continue

            babs2_good = chunk[(chunk['B-abs2'] != -99) & (~chunk['B-abs2'].isna())]['B-abs2']
            babs6_good = chunk[(chunk['B-abs6'] != -99) & (~chunk['B-abs6'].isna())]['B-abs6']
            if babs2_good.empty or babs6_good.empty:
                start_ = chunk_end
                continue

            found_3day = True
            logging.info(f"Calculating BrC R² for chunk: {start_.date()} to {chunk_end.date()}")
            steps = []
            r2_vals = []
            for i in range(61):
                stp = i / 10.0
                brc_sec = babs2_good - stp * babs6_good
                corr_ = np.corrcoef(brc_sec, babs6_good)[0, 1] if len(brc_sec) > 1 else 0.0
                r2_ = corr_ ** 2
                steps.append(stp)
                r2_vals.append(r2_)
            r2_vals = np.array(r2_vals)
            min_idx = np.argmin(r2_vals)
            min_r2 = r2_vals[min_idx]
            min_step = steps[min_idx]

            plt.figure(figsize=(10, 6))
            plt.plot(steps, r2_vals, label='R-squared')
            plt.annotate(f'Min R-squared\nStep: {min_step}\nR-squared: {min_r2:.6f}',
                         xy=(min_step, min_r2),
                         xytext=(min_step + 0.5, min_r2 - 0.05),
                         arrowprops=dict(facecolor='black', arrowstyle='->'))
            plt.xlabel('Step')
            plt.ylabel('R-squared')
            plt.title(f"BrC-Abs-Sec vs. B-abs6 {start_.date()} to {chunk_end.date()}")
            plt.legend()

            plot_name = f"BrCAbsSec_vs_Babs6_{start_.date()}_{chunk_end.date()}.png"
            plot_path = os.path.join(self.rsquared_dir, plot_name)
            plt.savefig(plot_path)
            plt.close()
            logging.info(f"Saved BrC R² plot: {plot_path}")

            mask = (df['Date_and_Time'] >= start_) & (df['Date_and_Time'] < chunk_end)
            df.loc[mask, 'BrC-abs-Sec'] = df.loc[mask, 'B-abs2'] - min_step * df.loc[mask, 'B-abs6']

            start_ = chunk_end

        if not found_3day:
            logging.info("No chunk with 3 consecutive days of AE33 data. 'BrC-abs-Sec' => NaN.")
            df['BrC-abs-Sec'] = float('nan')

    def min_r2_calculation_for_SOC(self):
        df = self.TCA_AE_hourly
        min_dt = df['Date_and_Time'].min()
        max_dt = df['Date_and_Time'].max()

        found_3day = False
        start_ = min_dt
        while start_ < max_dt:
            chunk_end = start_ + timedelta(days=self.Time_Delta)
            chunk = df[(df['Date_and_Time'] >= start_) & (df['Date_and_Time'] < chunk_end)]

            if chunk['Date_and_Time'].dt.date.nunique() < 3:
                start_ = chunk_end
                continue

            oc_good = chunk[(chunk['OC'] != -99) & (~chunk['OC'].isna())]['OC']
            bc_good = chunk[(chunk['AE33_BC6'] != -99) & (~chunk['AE33_BC6'].isna())]['AE33_BC6']
            if oc_good.empty or bc_good.empty:
                start_ = chunk_end
                continue

            found_3day = True
            logging.info(f"Calculating SOC R² for chunk: {start_.date()} to {chunk_end.date()}")
            steps = []
            r2_vals = []
            for i in range(101):
                stp = i / 10.0
                soc_ = oc_good - stp * bc_good
                corr_ = np.corrcoef(soc_, bc_good)[0, 1] if len(soc_) > 1 else 0.0
                r2_ = corr_ ** 2
                steps.append(stp)
                r2_vals.append(r2_)
            r2_vals = np.array(r2_vals)
            min_idx = np.argmin(r2_vals)
            min_r2 = r2_vals[min_idx]
            min_step = steps[min_idx]

            plt.figure(figsize=(10, 6))
            plt.plot(steps, r2_vals, label='R-squared')
            plt.annotate(f'Minimum R-squared\nStep: {min_step}\nR-squared: {min_r2:.6f}',
                         xy=(min_step, min_r2),
                         xytext=(min_step + 1, min_r2 - 0.1),
                         arrowprops=dict(facecolor='black', arrowstyle='->'))
            plt.xlabel('Step')
            plt.ylabel('R-squared')
            plt.title(f"SOC vs. BC {start_.date()} to {chunk_end.date()}")
            plt.legend()

            plot_name = f"SOC_vs_BC_{start_.date()}_{chunk_end.date()}.png"
            plot_path = os.path.join(self.rsquared_dir, plot_name)
            plt.savefig(plot_path)
            plt.close()
            logging.info(f"Saved SOC R² plot: {plot_path}")

            mask = (df['Date_and_Time'] >= start_) & (df['Date_and_Time'] < chunk_end)
            df.loc[mask, 'SOC'] = df.loc[mask, 'OC'] - min_step * df.loc[mask, 'AE33_BC6']

            start_ = chunk_end

        if not found_3day:
            logging.info("No chunk with 3 consecutive days of TCA data. 'SOC' => NaN.")
            df['SOC'] = float('nan')

    def write_to_excel(self, df: pd.DataFrame, sheet_name: str):
        if os.path.exists(self.output_xlsx):
            wb_out = load_workbook(self.output_xlsx)
            if sheet_name in wb_out.sheetnames:
                ws_out = wb_out[sheet_name]
                for row in ws_out.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                ws_out = wb_out.create_sheet(title=sheet_name)
        else:
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = sheet_name

        headers = df.columns.tolist()
        for col_idx, col_name in enumerate(headers, start=1):
            ws_out.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)

        for row_idx, row_vals in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_vals, start=1):
                ws_out.cell(row=row_idx, column=col_idx, value=val)

        ws_out.column_dimensions[get_column_letter(1)].width = 25

        wb_out.save(self.output_xlsx)
        wb_out.close()
        logging.info(f"Wrote sheet '{sheet_name}' with {len(df)} rows to {self.output_xlsx}")

    def write_constants_sheet(self):
        if os.path.exists(self.output_xlsx):
            wb_out = load_workbook(self.output_xlsx)
            if "constants" in wb_out.sheetnames:
                ws_const = wb_out["constants"]
                for row in ws_const.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                ws_const = wb_out.create_sheet("constants")
        else:
            wb_out = Workbook()
            ws_const = wb_out.active
            ws_const.title = "constants"

        ws_const.cell(row=1, column=1, value="Constant").font = Font(bold=True)
        ws_const.cell(row=1, column=2, value="Value").font = Font(bold=True)

        constants_data = [
            ("BC1", self.bc_multipliers["BC1"] * 1000),
            ("BC2", self.bc_multipliers["BC2"] * 1000),
            ("BC3", self.bc_multipliers["BC3"] * 1000),
            ("BC4", self.bc_multipliers["BC4"] * 1000),
            ("BC5", self.bc_multipliers["BC5"] * 1000),
            ("BC6", self.bc_multipliers["BC6"] * 1000),
            ("BC7", self.bc_multipliers["BC7"] * 1000),
            ("AAE_bb", self.AAE_bb),
            ("AAE_ff", self.AAE_ff),
            ("AAE_bc", self.AAE_bc),
            ("MAC_bb", self.MAC_bb),
            ("MAC_ff", self.MAC_ff),
            ("POA_POC_Ratio", self.POA_POC_Ratio),
            ("SOA_SOC_Ratio", self.SOA_SOC_Ratio),
            ("MAC_BrC_Prim", self.MAC_BrC_Prim),
            ("MAC_BrC_Sec", self.MAC_BrC_Sec),
            ("Time_Delta", self.Time_Delta),
            ("Start Date", self.start_date.strftime("%Y-%m-%d")),
            ("End Date", self.end_date.strftime("%Y-%m-%d")),
            ("Time Resolution (mins)", self.avg_interval_seconds // 60)
        ]

        row_idx = 2
        for k, v in constants_data:
            ws_const.cell(row=row_idx, column=1, value=k)
            ws_const.cell(row=row_idx, column=2, value=v)
            row_idx += 1

        wb_out.save(self.output_xlsx)
        wb_out.close()
        logging.info(f"Wrote constants sheet to {self.output_xlsx}")

    def produce_main_plots(self, output_prefix: str):
        if self.TCA_AE_hourly.empty:
            logging.warning("No data available to produce main plots.")
            return

        if not pd.api.types.is_datetime64_any_dtype(self.TCA_AE_hourly['Date_and_Time']):
            self.TCA_AE_hourly['Date_and_Time'] = pd.to_datetime(self.TCA_AE_hourly['Date_and_Time'])

        valid_df_ff = self.TCA_AE_hourly[self.TCA_AE_hourly['BC-ff'] != -99][['Date_and_Time', 'BC-ff']].dropna()
        valid_df_bb = self.TCA_AE_hourly[self.TCA_AE_hourly['BC-bb'] != -99][['Date_and_Time', 'BC-bb']].dropna()

        logging.info("Generating time-series (monthly summary) plot for BC-ff and BC-bb.")
        plt.figure(figsize=(12, 5))
        plt.plot(valid_df_ff['Date_and_Time'], valid_df_ff['BC-ff'], linestyle='-', alpha=0.7, label='BC-ff')
        plt.plot(valid_df_bb['Date_and_Time'], valid_df_bb['BC-bb'], linestyle='-', alpha=0.7, label='BC-bb')
        plt.xlabel("Date")
        plt.ylabel("BC (ng/m³)")
        plt.title(f"Monthly Summary: BC-ff & BC-bb\n{self.start_date.date()} to {self.end_date.date()}")
        plt.grid(True)
        plt.legend()
        ax = plt.gca()
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        plt.xticks(rotation=45)
        plt.tight_layout()
        ts_plot_path = os.path.join(self.plot_dir, f"{output_prefix}_BCff_BCbb_{self.start_date.date()}_{self.end_date.date()}.png")
        plt.savefig(ts_plot_path)
        plt.close()
        logging.info(f"Saved time-series plot: {ts_plot_path}")

        logging.info("Generating diurnal (24-hour) plot for BC-ff and BC-bb.")
        df_diurnal = self.TCA_AE_hourly.copy()
        df_diurnal['Hour'] = df_diurnal['Date_and_Time'].dt.hour
        df_diurnal = df_diurnal[(df_diurnal['BC-ff'] != -99) & (df_diurnal['BC-bb'] != -99)]
        if df_diurnal.empty:
            logging.warning("No valid data for diurnal plot.")
            return

        diurnal_stats = df_diurnal.groupby('Hour').agg(
            BCff_mean=('BC-ff', 'mean'),
            BCff_std=('BC-ff', 'std'),
            BCbb_mean=('BC-bb', 'mean'),
            BCbb_std=('BC-bb', 'std')
        ).reset_index()

        plt.figure(figsize=(10, 5))
        plt.plot(diurnal_stats['Hour'], diurnal_stats['BCff_mean'], '-', label='BC-ff Mean')
        plt.fill_between(diurnal_stats['Hour'],
                         diurnal_stats['BCff_mean'] - diurnal_stats['BCff_std'],
                         diurnal_stats['BCff_mean'] + diurnal_stats['BCff_std'],
                         alpha=0.2)
        plt.plot(diurnal_stats['Hour'], diurnal_stats['BCbb_mean'], '-', label='BC-bb Mean')
        plt.fill_between(diurnal_stats['Hour'],
                         diurnal_stats['BCbb_mean'] - diurnal_stats['BCbb_std'],
                         diurnal_stats['BCbb_mean'] + diurnal_stats['BCbb_std'],
                         alpha=0.2)
        plt.xlabel("Hour of Day")
        plt.ylabel("BC (ng/m³)")
        plt.title(f"Diurnal Plot: BC-ff & BC-bb\n{self.start_date.date()} to {self.end_date.date()}")
        plt.xticks(range(0,24))
        plt.grid(True)
        plt.legend()
        plt.tight_layout()
        diurnal_plot_path = os.path.join(self.plot_dir, f"{output_prefix}_Diurnal_{self.start_date.date()}_{self.end_date.date()}.png")
        plt.savefig(diurnal_plot_path)
        plt.close()
        logging.info(f"Saved diurnal plot: {diurnal_plot_path}")

def main():
    parser = argparse.ArgumentParser(description="CASS Speciation DB script with multiple sheets, prompting user for dates.")
    parser.add_argument("--const", type=str, default="../conf/constants.conf", help="Path to constants.conf")
    parser.add_argument("--dbconf", type=str, default="../conf/db.conf", help="Path to db.conf")
    parser.add_argument("--avg_interval", type=str, default="60mins",
                        help="Averaging interval. Options: 20mins, 30mins, 60mins, 120mins")
    args = parser.parse_args()

    allowed_intervals = {"20mins": 1200, "30mins": 1800, "60mins": 3600, "120mins": 7200}
    if args.avg_interval not in allowed_intervals:
        print("Invalid averaging interval. Allowed options: 20mins, 30mins, 60mins, 120mins.")
        sys.exit(1)
    avg_interval_seconds = allowed_intervals[args.avg_interval]

    # 1) Load configs
    constants_cfg = load_config_file(args.const)
    db_cfg = load_config_file(args.dbconf)

    # 2) Connect to DB and determine min/max dates, row counts, and time resolutions
    db_path = db_cfg["dbPath"]
    ae33_table = db_cfg.get("AE33_Table", "AE33_raw")
    tca_table = db_cfg.get("TCA_Table", "TCA_raw")

    cnx = sqlite3.connect(db_path)
    cursor = cnx.cursor()

    # For AE33_raw:
    cursor.execute(f"SELECT MIN(datetime), MAX(datetime), COUNT(*) FROM {ae33_table}")
    row = cursor.fetchone()
    ae33_min = row[0]
    ae33_max = row[1]
    ae33_count = row[2]

    cursor.execute(f"SELECT datetime FROM {ae33_table} ORDER BY datetime LIMIT 100")
    ae33_sample = [r[0] for r in cursor.fetchall()]
    from collections import Counter
    if len(ae33_sample) > 1:
        ae33_sample = pd.to_datetime(ae33_sample)
        ae33_diffs = [(ae33_sample[i+1] - ae33_sample[i]).total_seconds() / 60 for i in range(len(ae33_sample)-1)]
        ae33_mode = Counter(ae33_diffs).most_common(1)[0][0]
    else:
        ae33_mode = None

    # For TCA_raw:
    cursor.execute(f"SELECT MIN(StartTimeLocal), MAX(StartTimeLocal), COUNT(*) FROM {tca_table}")
    row = cursor.fetchone()
    tca_min = row[0]
    tca_max = row[1]
    tca_count = row[2]

    cursor.execute(f"SELECT StartTimeLocal FROM {tca_table} ORDER BY StartTimeLocal LIMIT 100")
    tca_sample = [r[0] for r in cursor.fetchall()]
    if len(tca_sample) > 1:
        tca_sample = pd.to_datetime(tca_sample)
        tca_diffs = [(tca_sample[i+1] - tca_sample[i]).total_seconds() / 60 for i in range(len(tca_sample)-1)]
        tca_mode = Counter(tca_diffs).most_common(1)[0][0]
    else:
        tca_mode = None

    cursor.close()
    cnx.close()

    print(f"Available date range for table '{ae33_table}': {pd.to_datetime(ae33_min).date()} to {pd.to_datetime(ae33_max).date()} ({ae33_count} records).")
    if ae33_mode is not None:
        print(f"Time resolution for {ae33_table}: {ae33_mode:.0f} minutes.")
    print(f"Available date range for table '{tca_table}': {pd.to_datetime(tca_min).date()} to {pd.to_datetime(tca_max).date()} ({tca_count} records).")
    if tca_mode is not None:
        print(f"Time resolution for {tca_table}: {tca_mode:.0f} minutes.")
    print()

    global_start = max(pd.to_datetime(ae33_min), pd.to_datetime(tca_min))
    global_end = min(pd.to_datetime(ae33_max), pd.to_datetime(tca_max))

    if global_start > global_end:
        print("ERROR: No overlapping date range between the tables. Exiting.")
        sys.exit(1)

    print(f"The combined overlap is from {global_start.date()} to {global_end.date()}.\n")

    # 4) Prompt the user for the date range to analyze.
    while True:
        user_start = input("Enter custom start date (YYYY-MM-DD) or type 'exit' to cancel: ").strip()
        user_end = input("Enter custom end date   (YYYY-MM-DD) or type 'exit' to cancel: ").strip()

        if user_start.lower() == "exit" or user_end.lower() == "exit":
            print("Exiting without performing analysis.")
            sys.exit(0)

        try:
            start_dt = datetime.strptime(user_start, "%Y-%m-%d")
            end_dt = datetime.strptime(user_end, "%Y-%m-%d")
        except ValueError:
            print("Invalid date format. Please try again.\n")
            continue

        if start_dt < global_start or end_dt > global_end:
            print("There is not enough data to process that range.\nPlease pick a range inside the intersection.\n")
            continue
        if end_dt < start_dt:
            print("End date cannot be before start date.\n")
            continue

        break

    # 5) Prompt for the averaging interval for analysis.
    while True:
        user_interval = input("Enter averaging interval ('20' minutes, '30' minutes, '60' minutes, '120' minutes) or type 'exit' to cancel: ").strip()
        if user_interval.lower() == "exit":
            print("Exiting without performing analysis.")
            sys.exit(0)
        allowed_intervals_prompt = {"20": 1200, "30": 1800, "60": 3600, "120": 7200}
        if user_interval in allowed_intervals_prompt:
            avg_interval_seconds = allowed_intervals_prompt[user_interval]
            break
        else:
            print("Invalid averaging interval. Please choose from 20, 30, 60, or 120.")

    run_start_time = datetime.now()
    runner = CassSpeciation(
        start_date=start_dt,
        end_date=end_dt,
        constants_cfg=constants_cfg,
        db_cfg=db_cfg,
        avg_interval_seconds=avg_interval_seconds
    )
    runner.run()

    run_end_time = datetime.now()
    total_run_time = run_end_time - run_start_time
    logging.info(f"Total run time: {total_run_time}")

    input("Processing complete. Press Enter to continue...")

if __name__ == "__main__":
    main()